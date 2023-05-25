import os
from openpyxl import Workbook, load_workbook

from pathlib import Path
from PyQt5.QtCore import QDir, QFile, QIODevice
from qgis._core import QgsMessageLog
from collections import OrderedDict

from stdm.ui.mobile_data_provider.mobile_provider_entity_reader import EntityReader

HOME = QDir.home().path()
DOCEXTENSION = '.xlsx'
STDM_FOLDER_PATH = '/.stdm'
PROVIDER_HOME = HOME + STDM_FOLDER_PATH + '/provider_forms/forms'


class XLSFormWriter:

    def __init__(self, file_name):
        self.file_name = file_name+DOCEXTENSION
        self.form = None

    def create_xls_form(self):

        Path(PROVIDER_HOME).mkdir(parents=True, exist_ok=True)
        home_directory = os.path.expanduser('~')
        self.form = os.path.join(home_directory, '.stdm', 'provider_forms', 'forms', self.file_name)
        failed = False
        sheet = None
        QgsMessageLog.logMessage("Form file path {path}".format(path=self.form), "STDM dev")

        # self.form = QFile(os.path.join(PROVIDER_HOME, self.file_name))

        # if not self.form.open(QIODevice.ReadWrite | QIODevice.Truncate |
        #                       QIODevice.Text):
        #     error = self.form.error()
        #     failed = False
        #     reason = ""
        #
        #     if error == QFile.OpenError:
        #         reason = "The file could not be opened."
        #
        #     if error == QFile.AbortError:
        #         reason = "The operation was aborted."
        #
        #     if error == QFile.TimeOutError:
        #         reason = "A timeout occurred."
        #
        #     if error == QFile.UnspecifiedError:
        #         reason = "An unspecified error occurred."
        #
        #     self.form = None
        #     return failed, reason
        #
        # else:
        #     # Load an existing workbook
        #     work_book = open(r'C:\Users\Lenovo\.stdm\provider_forms\forms\Informal_Settlement.xlsx', 'rb')
        # try:
        #     # Open or create the Excel file
        #     if os.path.exists(self.form):
        #         workbook = openpyxl.load_workbook(self.form)
        #     else:
        #         workbook = openpyxl.Workbook()
        #
        #     # Check if the worksheet already exists
        #     if 'survey' in workbook.sheetnames:
        #         worksheet1 = workbook['survey']
        #     else:
        #         # Create a new worksheet with name "MySheet"
        #         worksheet1 = workbook.create_sheet("survey")
        #
        #     # Add column headers to the new worksheet
        #     if 'type' not in [cell.value for cell in worksheet1[1]]:
        #         worksheet1['A'] = 'type'
        #     if 'name' not in [cell.value for cell in worksheet1[1]]:
        #         worksheet1['B'] = 'name'
        #     if 'label' not in [cell.value for cell in worksheet1[1]]:
        #         worksheet1['C'] = 'label'
        #
        #     if 'choices' in workbook.sheetnames:
        #         worksheet2 = workbook['choices']
        #     else:
        #         # Create a new worksheet with name "MySheet"
        #         worksheet2 = workbook.create_sheet("choices")
        #
        #     # Add column headers to the new worksheet
        #     if 'list_name' not in [cell.value for cell in worksheet2[2]]:
        #         worksheet2['A'] = 'list_name'
        #     if 'name' not in [cell.value for cell in worksheet2[2]]:
        #         worksheet2['B'] = 'name'
        #     if 'label' not in [cell.value for cell in worksheet2[2]]:
        #         worksheet2['C'] = 'label'
        #
        #     if 'settings' in workbook.sheetnames:
        #         worksheet3 = workbook['settings']
        #     else:
        #         # Create a new worksheet with name "MySheet"
        #         worksheet3 = workbook.create_sheet("settings")
        #
        #     # Add column headers to the new worksheet
        #     if 'form_title' not in [cell.value for cell in worksheet3[3]]:
        #         worksheet3['A'] = 'form_title'
        #     if 'form_id' not in [cell.value for cell in worksheet3[3]]:
        #         worksheet3['B'] = 'form_id'
        #     if 'version' not in [cell.value for cell in worksheet3[3]]:
        #         worksheet3['C'] = 'version'
        #     if 'instance_name' not in [cell.value for cell in worksheet3[3]]:
        #         worksheet3['D'] = 'instance_name'
        #
        #     # Save the workbook to a file
        #     workbook.save(self.form)
        #
        #     return True, ""
        #
        # except Exception as e:
        #     reason = "Error: {error}".format(error=str(e))
        #     return failed, reason
        sheets = {'survey': ['type', 'name', 'label'],
                  'choices': ['list_name', 'name', 'label'],
                  'settings': ['form_title', 'form_id', 'version', 'instance_name']}

        # try:
        #     # check if file exists
        #     if os.path.isfile(self.form):
        #         # file exists, load workbook
        #         workbook = load_workbook(self.form)
        #         # delete default first sheet if it exists
        #         if 'Sheet' in workbook.sheetnames:
        #             sheet = workbook['Sheet']
        #             workbook.remove(sheet)
        #             QgsMessageLog.logMessage("Default sheet 'Sheet' removed from the workbook.", "STDM dev")
        #         for sheetname, headers in sheets.items():
        #             # check if sheet exists
        #             if sheetname in workbook.sheetnames:
        #                 sheet = workbook[sheetname]
        #                 QgsMessageLog.logMessage(f"{sheetname} already exists in {self.form}.", "STDM dev")
        #             else:
        #                 # create new sheet
        #                 sheet = workbook.create_sheet(sheetname)
        #                 # add column headers
        #                 sheet.append(headers)
        #                 # save the workbook
        #                 workbook.save(self.form)
        #                 QgsMessageLog.logMessage(f"{sheetname} created in {self.form} with column headers: "
        #                                          f"{', '.join(headers)}.", "STDM dev")
        #             # check if headers exist
        #             existing_headers = [cell.value for cell in sheet[1]]
        #             if existing_headers == headers:
        #                 QgsMessageLog.logMessage(f"Column headers already exist in {sheetname}.", "STDM dev")
        #             else:
        #                 # add missing column headers
        #                 for header in headers:
        #                     if header not in existing_headers:
        #                         sheet.insert_cols(1)
        #                         sheet.cell(row=1, column=1, value=header)
        #                 # save the workbook
        #                 workbook.save(self.form)
        #                 QgsMessageLog.logMessage(f"Column headers added to {sheetname}.", "STDM dev")
        #     else:
        #         # create new workbook with sheets and column headers
        #         workbook = Workbook()
        #         for sheetname, headers in sheets.items():
        #             sheet = workbook.create_sheet(sheetname)
        #             sheet.append(headers)
        #         # delete default first sheet if it exists
        #         if 'Sheet' in workbook.sheetnames:
        #             sheet = workbook['Sheet']
        #             workbook.remove(sheet)
        #             QgsMessageLog.logMessage("Default sheet 'Sheet' removed from the workbook.", "STDM dev")
        #         # save the workbook
        #         workbook.save(self.form)
        #         QgsMessageLog.logMessage(f"{self.form} created with sheets: "
        #                                  f"{', '.join(sheets.keys())} and column headers.", "STDM dev")
        #     return True, ""
        #
        # except Exception as e:
        #     QgsMessageLog.logMessage(f"An error occurred: {e}", "STDM dev")
        #     reason = f"An error occurred: {e}"
        #     return failed, reason

        try:
            # check if file exists
            if os.path.isfile(self.form):
                # file exists, load workbook
                workbook = load_workbook(self.form)
                # delete default first sheet
                workbook.remove(workbook.active)
                for sheetname, headers in sheets.items():
                    # check if sheet exists
                    if sheetname in workbook.sheetnames:
                        sheet = workbook[sheetname]
                        QgsMessageLog.logMessage(f"{sheetname} already exists in {self.form}.", "STDM dev")
                    else:
                        # create new sheet
                        sheet = workbook.create_sheet(sheetname)
                        # add column headers
                        sheet.append(headers)
                        # save the workbook
                        workbook.save(self.form)
                        QgsMessageLog.logMessage(f"{sheetname} created in {self.form} with column headers: "
                                                 f"{', '.join(headers)}.", "STDM dev")
                    # check if headers exist
                    if sheet.max_row == 1:
                        existing_headers = [cell.value for cell in sheet[1]]
                        if existing_headers == headers:
                            QgsMessageLog.logMessage(f"Column headers already exist in {sheetname}.", "STDM dev")
                        else:
                            # add missing column headers
                            for header in headers:
                                if header not in existing_headers:
                                    sheet.insert_cols(1)
                                    sheet.cell(row=1, column=1, value=header)
                            # save the workbook
                            workbook.save(self.form)
                            QgsMessageLog.logMessage(f"Column headers added to {sheetname}.", "STDM dev")
                    else:
                        QgsMessageLog.logMessage(f"Cannot check headers in {sheetname} as it already has data.",
                                                 "STDM dev")
            else:
                # create new workbook with sheets and column headers
                workbook = Workbook()
                # delete default first sheet
                workbook.remove(workbook.active)
                for sheetname, headers in sheets.items():
                    sheet = workbook.create_sheet(sheetname)
                    sheet.append(headers)
                # save the workbook
                workbook.save(self.form)
                QgsMessageLog.logMessage(f"{self.form} created with sheets: "
                                         f"{', '.join(sheets.keys())} and column headers.", "STDM dev")
            return True, ""
        except Exception as e:
            QgsMessageLog.logMessage(f"An error occurred: {e}", "STDM dev")
            return failed, f"An error occurred: {e}"


class ProviderWriter(XLSFormWriter):

    def __init__(self, stdm_profile, entities):
        self.entities = entities
        self.profile = stdm_profile
        self.provider_file_name = self.profile.name.replace(' ', '_')
        QgsMessageLog.logMessage("STDM profile name {profile}".format(profile=self.provider_file_name),
                                 "STDM dev")

        XLSFormWriter.__init__(self, self.provider_file_name)

    def create_xsl_file(self) -> tuple[bool, str]:
        return self.create_xls_form()

    def write_data_to_xlsform(self) -> tuple[bool, str]:

        entities = OrderedDict()

        if self.form is None:
            QgsMessageLog.logMessage("Error: No file to write data!", "STDM dev")
            return False, "Error: No file to write data!"

        for entity_name in self.entities:
            reader = self.create_entity_reader(entity_name)
            entity_values = reader.read_attributes()
            for col, col_type in entity_values.items():
                if reader.on_column_info(col):
                    col_attributes = reader.format_lookup_items(col)
                    entities[col] = {col_type: col_attributes}
                else:
                    entities[col] = col_type

        row_data = {'survey': [],
                    'choices': []}

        for col, val in entities.items():

            # QgsMessageLog.logMessage("Entities col={col}, val={val}".format(col=col, val=val), "STDM dev")

            if val == 'VARCHAR':
                row = OrderedDict()
                row['type'] = 'text'
                row['name'] = col
                row_data['survey'].append(row)

            elif val == 'INT' or val == 'DOUBLE' or val == 'AUTO_GENERATED':
                row = OrderedDict()
                row['type'] = 'integer'
                row['name'] = col
                row_data['survey'].append(row)

            elif val == 'DATE':
                row = OrderedDict()
                row['type'] = 'date'
                row['name'] = col
                row_data['survey'].append(row)

            elif val == 'GEOMETRY':
                row = OrderedDict()
                row['type'] = 'geoshape'
                row['name'] = col
                row_data['survey'].append(row)

            elif isinstance(val, dict):
                row = OrderedDict()

                data_type = list(val)[0]
                # QgsMessageLog.logMessage(f"{data_type}", "STDM dev")
                if data_type == 'LOOKUP':
                    row['type'] = 'select_one {col}'.format(col=col)
                    row['name'] = col
                    row_data['survey'].append(row)
                elif data_type == 'MULTIPLE_SELECT':
                    row['type'] = 'select_multiple {col}'.format(col=col)
                    row['name'] = col
                    row_data['survey'].append(row)
                elif data_type == 'ADMIN_SPATIAL_UNIT':
                    row['type'] = 'select_multiple {col}'.format(col=col)
                    row['name'] = col
                    row_data['survey'].append(row)

                for v in val.values():
                    row['list_name'] = col
                    row['name'] = v
                    row_data['choices'].append(row)

        QgsMessageLog.logMessage("Rows: rows={rows}".format(rows=row_data), "STDM dev")

        # load the workbook
        workbook = load_workbook(self.form)

        # delete default first sheet
        sheets = {'survey': ['type', 'name', 'label'],
                  'choices': ['list_name', 'name', 'label'],
                  'settings': ['form_title', 'form_id', 'version', 'instance_name']}

        if 'Sheet' in workbook.sheetnames:
            sheet = workbook['Sheet']
            workbook.remove(sheet)
            print("Default sheet 'Sheet' removed from the workbook.")

        for sheetname, headers in sheets.items():
            # check if sheet exists
            if sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                print(f"{sheetname} already exists in {self.form}.")
            else:
                print(f"{sheetname} does not exist in {self.form}.")
                continue

            # add missing column headers
            existing_headers = [cell.value for cell in sheet[1]]
            for header in headers:
                if header not in existing_headers:
                    sheet.insert_cols(1)
                    sheet.cell(row=1, column=1, value=header)
                    print(f"{header} added to {sheetname}.")

            # add data to the sheet
            for _row_data in row_data.get(sheetname, []):
                current_row = []

                QgsMessageLog.logMessage("Headers: {h} ".format(h=headers), "STDM dev")
                row = []
                flag = False
                for header in headers:
                    QgsMessageLog.logMessage("Raw values: {r} header {h}".format(r=_row_data, h=header), "STDM dev")
                    column_numbers = [i for i, x in enumerate(existing_headers) if x == header]
                    # if len(column_numbers) == 1:

                    QgsMessageLog.logMessage("In row data: {d}".format(d=_row_data),
                                             "STDM dev")
                    for type, data in _row_data.items():
                        QgsMessageLog.logMessage("Processing dict: {d}, {v}".format(d=type, v=data),
                                                 "STDM dev")
                        if isinstance(data, dict):
                            # if type == 'type':
                            #     row.append(data)
                            #     sheet.append(row)
                            for v in data.keys():
                                choice_row = ['name', v]
                                QgsMessageLog.logMessage("Raw values after looping: {r} header {h}"
                                                         .format(r=choice_row, h=header), "STDM dev")
                                sheet.append(choice_row)
                                sheet.cell(row=len(sheet['A']) + 1, column=column_numbers[0] + 1,
                                           value=row_data.get(header))
                                QgsMessageLog.logMessage("Raw data: {d}".format(d=choice_row),
                                                         "STDM dev")
                            flag = True
                        else:
                            if header == type:
                                row.append(data)
                    QgsMessageLog.logMessage("Raw data: {d}".format(d=row),
                                             "STDM dev")

                if not flag:
                    sheet.append(row)
                    # sheet.cell(row=len(sheet['A']) + 1, column=column_numbers[0] + 1,
                    #            value=row_data.get(header))

                # else:
                #     row = []
                #     flag = False
                #     QgsMessageLog.logMessage("In row data: {d}".format(d=_row_data),
                #                              "STDM dev")
                #     for type, data in _row_data.items():
                #         QgsMessageLog.logMessage("Processing dict: {d}, {v}".format(d=type, v=data),
                #                                  "STDM dev")
                #         if isinstance(data, dict):
                #             for v in data.keys():
                #                 choice_row = ['name', v]
                #                 QgsMessageLog.logMessage("Raw values after looping: {r} header {h}"
                #                                          .format(r=choice_row, h=header), "STDM dev")
                #                 sheet.append(choice_row)
                #                 sheet.cell(row=len(sheet['A'])+1, column=column_numbers[-1]+1, value=row_data.get(header))
                #                 QgsMessageLog.logMessage("Raw data: {d}".format(d=choice_row),
                #                                          "STDM dev")
                #             flag = True
                #         else:
                #             row.append(row.append(data))
                #     QgsMessageLog.logMessage("Raw data: {d}".format(d=row),
                #                              "STDM dev")
                #     if not flag:
                #         sheet.append(row)
                #         sheet.cell(row=len(sheet['A'])+1, column=column_numbers[-1]+1, value=row_data.get(header))

                # QgsMessageLog.logMessage("Current row {r}".format(r=current_row), "STDM dev")

                    # if isinstance(_row_data, dict):
                    #     for k, v in _row_data.items():
                    #         row.append(v)
                    #     # row.append(_row_data.get(header))
                    #     sheet.append(row)
                    # else:
                    #     sheet.append(row)
            print(f"Example data added to {sheetname}.")

        # save the workbook
        workbook.save(self.form)
        print(f"Data saved to {self.form}.")

        return True, "File found"

    def create_entity_reader(self, entity):
        """
        Initialize the reader class after each entity to avoid
        redundant data
        """
        self.entity_read = EntityReader(entity)
        return self.entity_read
